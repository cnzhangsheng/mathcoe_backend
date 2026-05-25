"""
Batch import questions from Excel + ZIP
"""
import io
import logging
import os
import re
import uuid
import zipfile
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.question import Question
from app.utils.id_generator import short_id

logger = logging.getLogger(__name__)

IMG_PLACEHOLDER_RE = re.compile(r"\{img\}")


class ImportError:
    """Represents a single row-level import failure."""

    def __init__(self, row: int, message: str):
        self.row = row
        self.message = message

    def to_dict(self) -> dict:
        return {"row": self.row, "message": self.message}


class ImportResult:
    """Aggregated result of a batch import."""

    def __init__(self):
        self.total = 0
        self.imported = 0
        self.errors: list[ImportError] = []

    def to_dict(self) -> dict:
        return {
            "total": self.total,
            "imported": self.imported,
            "failed": len(self.errors),
            "errors": [e.to_dict() for e in self.errors],
        }


def _save_uploaded_image(file_bytes: bytes, original_name: str) -> str:
    """Save an image to the uploads directory and return its absolute URL."""
    year_month = datetime.now().strftime("%Y%m")
    upload_dir = os.path.join("app/static/uploads", year_month)
    os.makedirs(upload_dir, exist_ok=True)
    ext = original_name.rsplit(".", 1)[-1] if "." in original_name else "png"
    filename = f"{datetime.now().strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}.{ext}"
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, "wb") as f:
        f.write(file_bytes)
    base = settings.server_host.rstrip("/")
    return f"{base}/api/v1/static/uploads/{year_month}/{filename}"


def _parse_zip(zip_bytes: bytes) -> dict[int, dict]:
    """Parse ZIP file and build image mapping.

    Returns:
        {question_number: {"content": [url], "options": {"a": [url], "b": [url], ...}}}
    """
    mapping: dict[int, dict] = {}

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            # Skip directories, macOS __MACOSX metadata, and Apple Double files
            if not name or name.endswith("/"):
                continue
            if name.startswith("__MACOSX/") or name.startswith("."):
                continue
            basename = os.path.basename(name)
            if basename.startswith("._") or basename.startswith("."):
                continue
            if not basename.lower().endswith((".png", ".jpg", ".jpeg")):
                continue

            stem = os.path.splitext(basename)[0]

            # Parse: "3" -> content image for question 3
            # Parse: "3a" or "3A" -> option A image for question 3
            match = re.match(r"^(\d+)([a-eA-E])?$", stem)
            if not match:
                logger.debug("Skipping unmatched file in ZIP: %s", name)
                continue

            qnum = int(match.group(1))
            opt_label = match.group(2)

            data = zf.read(name)
            url = _save_uploaded_image(data, name)

            if qnum not in mapping:
                mapping[qnum] = {"content": [], "options": {}}

            if opt_label:
                opt_key = opt_label.lower()
                if opt_key not in mapping[qnum]["options"]:
                    mapping[qnum]["options"][opt_key] = []
                mapping[qnum]["options"][opt_key].append(url)
            else:
                mapping[qnum]["content"].append(url)

    return mapping


def _replace_img_placeholders(text: str, image_urls: list[str], img_style: str = "") -> str:
    """Replace {img} placeholders in text with <img> tags."""
    if not image_urls:
        return text
    style_attr = f' style="{img_style}"' if img_style else ""
    for url in image_urls:
        text = IMG_PLACEHOLDER_RE.sub(f'<img src="{url}"{style_attr}/>', text, 1)
    return text


def _strip_html(src: str) -> str:
    """Strip HTML tags for plain-text extraction."""
    return re.sub(r"<[^>]+>", "", src).strip()


def _extract_title_from_content(content_text: str) -> str:
    """Extract a short title from HTML content."""
    plain = _strip_html(content_text)
    return plain[:50] if plain else "批量导入题目"


async def batch_import(
    db: AsyncSession,
    excel_bytes: bytes,
    zip_bytes: bytes | None = None,
) -> ImportResult:
    """Execute batch import of questions."""
    result = ImportResult()

    # --- 1. Parse ZIP (if provided) ---
    image_map: dict[int, dict] = {}
    if zip_bytes:
        try:
            image_map = _parse_zip(zip_bytes)
        except Exception as e:
            logger.error("ZIP parse failed: %s", e)
            result.errors.append(ImportError(0, f"ZIP 文件解析失败: {e}"))
            return result

    # --- 2. Parse Excel ---
    try:
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        logger.info(
            "Excel parsed: sheet='%s', sheet_names=%s, total_rows=%d",
            ws.title, wb.sheetnames, len(rows),
        )
        if rows:
            logger.info("Row 0 (header): %s", rows[0])
            if len(rows) > 1:
                logger.info("Row 1 (first data): %s", rows[1])
            else:
                logger.warning("Only header row found in active sheet '%s'", ws.title)

        # If active sheet has only header, try other sheets
        if len(rows) < 2:
            for name in wb.sheetnames:
                if name == ws.title:
                    continue
                alt_ws = wb[name]
                alt_rows = list(alt_ws.iter_rows(values_only=True))
                logger.info("Sheet '%s': total_rows=%d", name, len(alt_rows))
                if alt_rows:
                    logger.info("  Row 0: %s", alt_rows[0])
                if len(alt_rows) >= 2:
                    ws = alt_ws
                    rows = alt_rows
                    logger.info("Using sheet '%s' instead", name)
                    break
    except Exception as e:
        logger.exception("Excel parse failed:")
        result.errors.append(ImportError(0, f"Excel 文件解析失败: {e}"))
        return result

    if not rows or len(rows) < 2:
        result.errors.append(ImportError(0, f"Excel 文件数据不足（共 {len(rows) if rows else 0} 行，需要至少 1 行数据）"))
        return result

    # Skip header row, process data rows
    data_rows = rows[1:]
    result.total = len(data_rows)

    questions_to_create: list[Question] = []
    seen_numbers: set[int] = set()

    for row_idx, row in enumerate(data_rows):
        lineno = row_idx + 2  # Excel row number (1-indexed, header is row 1)
        try:
            if not row or all(v is None for v in row):
                continue  # skip completely empty rows

            qnum = _safe_int(row[0])
            if qnum is None:
                result.errors.append(ImportError(lineno, "题号无效或为空"))
                continue

            if qnum in seen_numbers:
                result.errors.append(ImportError(lineno, f"题号 {qnum} 重复"))
                continue
            seen_numbers.add(qnum)

            qtype = _safe_str(row[1]).lower()
            if qtype not in ("single", "multiple"):
                result.errors.append(ImportError(lineno, f"题型 '{row[1]}' 无效，需为 single 或 multiple"))
                continue

            topic_id = _safe_int(row[2])
            if topic_id is None:
                result.errors.append(ImportError(lineno, "专题ID 无效或为空"))
                continue

            difficulty = _safe_int(row[3])
            if difficulty is None or difficulty < 1 or difficulty > 6:
                result.errors.append(ImportError(lineno, "难度需为 1-6 的整数"))
                continue

            content_text = _safe_str(row[4])
            if not content_text:
                result.errors.append(ImportError(lineno, "题目内容为空"))
                continue

            # Options (columns F-J, indices 5-9)
            raw_options = [_safe_str(row[i]) for i in range(5, 10)]

            answer = _safe_str(row[10])
            if not answer:
                result.errors.append(ImportError(lineno, "正确答案为空"))
                continue

            explanation_text = _safe_str(row[11])
            source_year = _safe_int(row[12])

            # --- 3. Process images for this question ---
            q_images = image_map.get(qnum, {"content": [], "options": {}})

            # Replace {img} in content (width: 100%)
            processed_content = _replace_img_placeholders(content_text, q_images.get("content", []), "width: 100%;")

            # Build options array
            option_labels = ["A", "B", "C", "D", "E"]
            options_list = []
            for opt_idx, opt_text in enumerate(raw_options):
                if opt_idx >= len(option_labels):
                    break
                label = option_labels[opt_idx]
                opt_key = label.lower()
                opt_img_urls = q_images.get("options", {}).get(opt_key, [])

                has_text = bool(opt_text and opt_text.strip())
                has_image = bool(opt_img_urls)

                if not has_text and not has_image:
                    # Empty column and no image — skip this option entirely
                    continue

                if has_text or has_image:
                    processed_opt_text = _replace_img_placeholders(opt_text or "", opt_img_urls, "width: 50%;")
                    options_list.append({
                        "label": label,
                        "text": processed_opt_text,
                        "format": "html",
                    })

            # Validate answer against available options
            answer_labels = {o["label"] for o in options_list}
            if qtype == "single":
                if answer not in answer_labels:
                    result.errors.append(ImportError(
                        lineno, f"答案 '{answer}' 超出选项范围 ({', '.join(sorted(answer_labels))})"
                    ))
                    continue
            else:
                # Multiple choice: answer like "A,B"
                parts = [a.strip() for a in answer.split(",")]
                if not all(p in answer_labels for p in parts):
                    result.errors.append(ImportError(
                        lineno, f"答案 '{answer}' 中存在超出选项范围的选项 ({', '.join(sorted(answer_labels))})"
                    ))
                    continue

            # Process explanation (width: 100%)
            processed_explanation = _replace_img_placeholders(explanation_text or "", q_images.get("content", []), "width: 100%;")

            # Build title
            title = _extract_title_from_content(content_text)

            # Create question object
            question = Question(
                id=short_id(),
                topic_id=topic_id,
                title=title,
                content={
                    "text": processed_content,
                    "format": "html",
                },
                question_type=qtype,
                options=options_list if options_list else None,
                answer=answer,
                explanation={
                    "text": processed_explanation,
                    "format": "html",
                } if (explanation_text or q_images.get("content")) else None,
                difficulty_level=difficulty,
                source_year=source_year,
                status="published",
            )
            questions_to_create.append(question)

        except Exception as e:
            logger.exception("Row %d processing error:", lineno)
            result.errors.append(ImportError(lineno, f"处理异常: {e}"))

    # --- 4. Bulk insert ---
    if questions_to_create:
        try:
            db.add_all(questions_to_create)
            await db.commit()
            result.imported = len(questions_to_create)
        except Exception as e:
            await db.rollback()
            logger.exception("Bulk insert failed:")
            result.errors.append(ImportError(0, f"批量写入数据库失败: {e}"))

    return result


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None
